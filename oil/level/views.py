from django.shortcuts import render
import numpy as np
from level.models import Stan4,Stan3,Stan2,Stan1
from sklearn.linear_model import LinearRegression
import copy

from datetime import datetime
from openpyxl import load_workbook, workbook
from os.path import join, abspath


def index(request):
    """
    # ------------------------------------Для БД-----------------------------
    # функция чтения из экселя аргументы: номер столбца, рабочая книга
    print(head)


    def read_x(col_number, source):
        name = []
        for col in source.iter_rows(min_row=2, min_col=col_number,
                                    max_col=col_number, max_row=source.max_row):
            for cell in col:
                name.insert(0, cell.value)
        return name


    Date = read_x(1, source=work_sheet)
    level = read_x(2, work_sheet)
    temp = read_x(7, work_sheet)
    power = read_x(8, work_sheet)
    pocket = read_x(10, work_sheet)

    for i in range(len(Date)):
        if 'января' in Date[i]:
            Date[i] = Date[i].replace('января', 'january')
        if 'февраля' in Date[i]:
            Date[i] = Date[i].replace('февраля', 'February')
        if 'марта' in Date[i]:
            Date[i] = Date[i].replace('марта', 'march')
        if 'апреля' in Date[i]:
            Date[i] = Date[i].replace('апреля', 'april')
        if 'мая' in Date[i]:
            Date[i] = Date[i].replace('мая', 'may')
        if 'июня' in Date[i]:
            Date[i] = Date[i].replace('июня', 'june')
        if 'июля' in Date[i]:
            Date[i] = Date[i].replace('июля', 'july')
        if 'августа' in Date[i]:
            Date[i] = Date[i].replace('августа', 'august')
        if 'сентябр' in Date[i]:
            Date[i] = Date[i].replace('сентября', 'september')
        if 'октября' in Date[i]:
            Date[i] = Date[i].replace('октября', 'october')
        if 'ноября' in Date[i]:
            Date[i] = Date[i].replace('ноября', 'november')
        if 'декабря' in Date[i]:
            Date[i] = Date[i].replace('декабря', 'december')
        Date[i] = datetime.strptime(Date[i], '%d %B %Y г., %H:%M')


    # формируем dataframe для работы с датами
    oil_df = pd.DataFrame({'Расход масла': level,
                           'Температура': temp,
                           'Питание': power,
                           '№ пакета': pocket}, index=Date)

    oil_df = oil_df.resample('5T').first()  # интервал в 5 минут с усреднением
    oil_df.ffill(inplace=True)  # заполнение отсутствующих измерений


    level = oil_df['Расход масла']
    level = level.tolist()

    temp = oil_df['Температура']
    temp = temp.tolist()

    power = oil_df['Питание']
    power = power.tolist()

    pocket = oil_df['№ пакета']
    pocket = pocket.tolist()

    data = oil_df.index
    data = data.tolist()


    for i in range(len(level)):
        b = Stan3(date_value=data[i], level_value=level[i],
                  temp_value=temp[i], package_value=pocket[i], power_value=power[i])
        b.save()

    """
    start_date='2020-08-01'
    end_date='2020-08-10'
    if 'start_date' in request.GET and request.GET['start_date']:
        start_date=request.GET['start_date']

    if 'end_date' in request.GET and request.GET['end_date']:
        end_date=request.GET['end_date']



    # формирование выборки

    def bd_query (stan,start_date, end_date):
        test=stan.objects.filter(date_value__range=[start_date, end_date])
        level=[]
        data=[]
        for i in range (len(test)):
            data.append(test[i].date_value)
            level.append(test[i].level_value)


        for i in range (len(data)):
            data[i]=str(data[i])[0:16]
        return (level, data)
    """ пример даты 2020-08-01"""
    level_1,data_1=bd_query(Stan1, start_date, end_date)
    level_2, data_2 = bd_query(Stan2, start_date, end_date)
    level_3, data_3 = bd_query(Stan3, start_date, end_date)
    level_4, data_4 = bd_query(Stan4, start_date, end_date)
    #функция построения модели

    def line_model(arg1, arg2=0):
        x = np.arange(arg2, (arg2 + (len(arg1))), 1)
        x = x.reshape((-1, 1))
        y = np.array(arg1)

        # рассчет параметров
        oil_model = LinearRegression().fit(x, y)
        return (oil_model.intercept_, float(oil_model.coef_))

    def corr(arg):
        res = []
        for i in range(0, len(arg)):
            if i == 0 or i == (len(arg)):
                res.append(arg[i])
            else:
                if abs(res[i - 1] - arg[i]) > 5 and (
                        abs(arg[i] - arg[i + 1]) > 5 or ((abs(arg[i] - arg[i + 5])) > 5) or (
                        (abs(arg[i] - arg[i + 10])) > 5)):
                    res.append(res[i - 1])
                else:
                    res.append(arg[i])
        return res

    # функция определения доливов
    def doliv(arg, bord):
        res = [0]
        for i in range(1, (len(arg))):
            if (arg[i] - arg[i - 1]) > bord:
                res.append(res[i - 1] + (arg[i] - arg[i - 1]))
            else:
                res.append(res[i - 1])

        return res
    # величина контрольной ошибки
    mis = 8

    cor1=corr(level_1)
    dol1 = doliv(cor1, 5)
    level_cor1 = []
    for i in range(len(cor1)):
        level_cor1.append(cor1[i] - dol1[i])


    coef = line_model(level_cor1, 0)
    predict1 = []
    speed_data1 = []

    for i in range(len(level_cor1)):
        predict1.append(coef[0] + coef[1] * (i + 1))
        speed_data1.append(-coef[1] * 12)
        if i < (len(level_cor1) - 50) and abs(level_cor1[i] - predict1[i]) > mis:
            coef = line_model(level_cor1[i:i + 50], i)

    cor2 = corr(level_2)
    dol2 = doliv(cor2, 5)
    level_cor2 = []
    for i in range(len(cor2)):
        level_cor2.append(cor2[i] - dol2[i])

    coef = line_model(level_cor2, 0)
    predict2 = []
    speed_data2 = []

    for i in range(len(level_cor2)):
        predict2.append(coef[0] + coef[1] * (i + 1))
        speed_data2.append(-coef[1] * 12)
        if i < (len(level_cor2) - 50) and abs(level_cor2[i] - predict2[i]) > mis:
            coef = line_model(level_cor2[i:i + 50], i)

    cor3 = corr(level_3)
    dol3 = doliv(cor3, 5)
    level_cor3 = []
    for i in range(len(cor3)):
        level_cor3.append(cor3[i] - dol3[i])

    coef = line_model(level_cor3, 0)
    predict3 = []
    speed_data3 = []

    for i in range(len(level_cor3)):
        predict3.append(coef[0] + coef[1] * (i + 1))
        speed_data3.append(-coef[1] * 12)
        if i < (len(level_cor3) - 50) and abs(level_cor3[i] - predict3[i]) > mis:
            coef = line_model(level_cor3[i:i + 50], i)

    cor4 = corr(level_4)
    dol4 = doliv(cor4, 5)
    level_cor4 = []
    for i in range(len(cor4)):
        level_cor4.append(cor4[i] - dol4[i])

    coef = line_model(level_cor4, 0)
    predict4 = []
    speed_data4 = []

    for i in range(len(level_cor4)):
        predict4.append(coef[0] + coef[1] * (i + 1))
        speed_data4.append(-coef[1] * 12)
        if i < (len(level_cor4) - 50) and abs(level_cor4[i] - predict4[i]) > mis:
            coef = line_model(level_cor4[i:i + 50], i)
    #подготовка данных о расходе и доливе, перевод в проценты

    consumption=[0,0,0,0]

    consumption[0]=level_cor1[0]-level_cor1[-1]
    consumption[1] = level_cor2[0] - level_cor2[-1]
    consumption[2] = level_cor3[0] - level_cor3[-1]
    consumption[3] = level_cor4[0] - level_cor4[-1]
    s_c=sum(consumption)
    con=copy.deepcopy(consumption)
    consumption[:] = [x / s_c for x in consumption]
    sum_dol=[dol1[-1], dol2[-1], dol3[-1], dol4[-1]]
    s_d=sum(sum_dol)
    dol=copy.deepcopy(sum_dol)
    sum_dol[:] = [x / s_d for x in sum_dol]
    print(con)
    context={"data1": data_1, "level1": level_1,
            "level2": level_2,
            "level3": level_3,
            "level4": level_4,
             "model1": predict1, "model2": predict2,
             "model3": predict3, "model4": predict4,
             'consumption_per': consumption, 'doliv_per':sum_dol,
             'consumption': con, 'doliv': dol,
             'speed1': speed_data1, 'speed2': speed_data2,
             'speed3': speed_data3,'speed4': speed_data4}


    return render(request, 'index.html', context)